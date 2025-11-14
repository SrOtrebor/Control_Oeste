import os
import re
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

from config import (
    BASE_DIR, REGISTROS_DIARIOS_DIR,
    EXCEL_FAP, EXCEL_FAO, EXCEL_EXCEPCIONES, EXCEL_NOMINAS,
    COL_DNI, COL_NOMBRE_APELLIDO, COL_NUM_PERMISO, COL_VENCE, COL_LOCAL, COL_TAREA, COL_TIPO_PERMISO,
    COL_DNI_FAP_ORIGINAL, COL_NOMBRE_FAP_ORIGINAL, COL_APELLIDO_FAP_ORIGINAL, COL_NUM_PERMISO_FAP_ORIGINAL, COL_VENCE_FAP_ORIGINAL, COL_LOCAL_FAP_ORIGINAL,
    COL_DNI_FAO_ORIGINAL, COL_NOMBRE_FAO_ORIGINAL, COL_APELLIDO_FAO_ORIGINAL, COL_NUM_PERMISO_FAO_ORIGINAL, COL_VENCE_FAO_ORIGINAL, COL_LOCAL_FAO_ORIGINAL, COL_TAREA_FAO_ORIGINAL
)

# --- VARIABLES GLOBALES PARA DATAFRAMES Y TIEMPOS DE MODIFICACIÓN ---
df_fap, df_fao, df_excepciones, df_nominas = pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), pd.DataFrame()
ult_mod_fap, ult_mod_fao, ult_mod_excepciones, ult_mod_nominas = 0, 0, 0, 0

def formatear_excel(nombre_archivo):
    """
    Aplica formato profesional a un archivo Excel: ajusta columnas,
    formatea el encabezado y congela la primera fila.
    """
    try:
        workbook = load_workbook(nombre_archivo)
        worksheet = workbook.active

        # Estilo para el encabezado
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_alignment = Alignment(horizontal='center', vertical='center')

        # Aplicar estilo al encabezado y ajustar ancho de columnas
        for col in worksheet.columns:
            max_length = 0
            column = col[0].column_letter # Obtener la letra de la columna
            
            # Ajustar el ancho de la columna
            for cell in col:
                if cell.coordinate in worksheet.merged_cells: # Ignorar celdas combinadas
                    continue
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column].width = adjusted_width

            # Formatear la celda del encabezado
            header_cell = worksheet[f"{column}1"]
            header_cell.font = header_font
            header_cell.fill = header_fill
            header_cell.alignment = header_alignment

        # Congelar la fila del encabezado
        worksheet.freeze_panes = 'A2'
        
        workbook.save(nombre_archivo)
        print(f"INFO: Formato aplicado correctamente a '{os.path.basename(nombre_archivo)}'.")

    except Exception as e:
        print(f"Error al aplicar formato a {os.path.basename(nombre_archivo)}: {e}")

# --- FUNCIÓN AUXILIAR PARA LIMPIAR DNI/CUIL ---
def extraer_dni_de_cuil(valor):
    """
    Toma un valor (que puede ser DNI o CUIL), lo limpia y extrae
    los 8 dígitos del DNI si es un CUIL.
    """
    valor_str = str(valor).strip().replace('-', '')
    
    if len(valor_str) == 11 and valor_str.isdigit():
        return valor_str[2:10] # Extrae los 8 dígitos del medio
    elif len(valor_str) >= 7 and valor_str.isdigit():
        return valor_str # Devuelve DNI de 7 u 8 dígitos
    else:
        return valor_str # Devuelve el valor original si no es un formato esperado

def cargar_y_procesar_excel(archivo_excel, ultima_modificacion, tipo_permiso, mapa_columnas, df_actual, header=0, dni_col_original=None):
    """
    Función genérica para cargar y procesar un archivo Excel, recargando solo si ha cambiado.
    Asegura que la columna de DNI se lea como texto y se normalice.
    """
    try:
        if not os.path.exists(archivo_excel):
            return pd.DataFrame(), 0
        mod_time = os.path.getmtime(archivo_excel)
        if mod_time == ultima_modificacion and not df_actual.empty:
            return df_actual, ultima_modificacion

        print(f"INFO: Detectado cambio en '{os.path.basename(archivo_excel)}'. Recargando...")
        
        dtype_map = {}
        if dni_col_original is not None:
            dtype_map[dni_col_original] = str

        df = pd.read_excel(archivo_excel, header=header, dtype=dtype_map)

        df.columns = [str(c).strip() for c in df.columns]
        df.rename(columns=mapa_columnas, inplace=True)

        if COL_DNI in df.columns:
            df[COL_DNI] = df[COL_DNI].astype(str).str.replace(r'\.0$', '', regex=True).str.replace(r'[\.\s-]', '', regex=True).str.strip()
            df.loc[df[COL_DNI].str.lower() == 'nan', COL_DNI] = ''
            df = df[df[COL_DNI] != ''].copy()
            df[COL_DNI] = df[COL_DNI].apply(extraer_dni_de_cuil)

        if 'Nombre' in df.columns and 'Apellido' in df.columns and COL_NOMBRE_APELLIDO not in df.columns:
            df[COL_NOMBRE_APELLIDO] = df['Nombre'].fillna('') + ' ' + df['Apellido'].fillna('')

        df[COL_TIPO_PERMISO] = tipo_permiso
        
        print(f"-> Recargado y procesado: {os.path.basename(archivo_excel)}")
        return df, mod_time

    except FileNotFoundError:
        print(f"ADVERTENCIA: No se encontró el archivo {archivo_excel}. Se continuará sin él.")
        return pd.DataFrame(), 0
    except Exception as e:
        print(f"Error Crítico al procesar {archivo_excel}: {e}")
        return df_actual, ultima_modificacion

def cargar_autorizaciones():
    global df_fap, ult_mod_fap, df_fao, ult_mod_fao, df_excepciones, ult_mod_excepciones, df_nominas, ult_mod_nominas
    
    mapa_cols_fap = {COL_DNI_FAP_ORIGINAL: COL_DNI, COL_NOMBRE_FAP_ORIGINAL: 'Nombre', COL_APELLIDO_FAP_ORIGINAL: 'Apellido', COL_NUM_PERMISO_FAP_ORIGINAL: COL_NUM_PERMISO, COL_VENCE_FAP_ORIGINAL: COL_VENCE, COL_LOCAL_FAP_ORIGINAL: COL_LOCAL}
    df_fap, ult_mod_fap = cargar_y_procesar_excel(EXCEL_FAP, ult_mod_fap, 'FAP', mapa_cols_fap, df_fap, header=1, dni_col_original=COL_DNI_FAP_ORIGINAL)
    if not df_fap.empty and COL_VENCE in df_fap.columns:
        df_fap[COL_VENCE] = pd.to_datetime(df_fap[COL_VENCE], format='%d/%m/%Y', errors='coerce')

    mapa_cols_fao = {COL_DNI_FAO_ORIGINAL: COL_DNI, COL_NOMBRE_FAO_ORIGINAL: 'Nombre', COL_APELLIDO_FAO_ORIGINAL: 'Apellido', COL_NUM_PERMISO_FAO_ORIGINAL: COL_NUM_PERMISO, COL_VENCE_FAO_ORIGINAL: COL_VENCE, COL_LOCAL_FAO_ORIGINAL: COL_LOCAL, COL_TAREA_FAO_ORIGINAL: COL_TAREA}
    df_fao, ult_mod_fao = cargar_y_procesar_excel(EXCEL_FAO, ult_mod_fao, 'FAO', mapa_cols_fao, df_fao, header=1, dni_col_original=COL_DNI_FAO_ORIGINAL)
    if not df_fao.empty and COL_VENCE in df_fao.columns:
        df_fao[COL_VENCE] = pd.to_datetime(df_fao[COL_VENCE], format='%d/%m/%Y', errors='coerce')
    
    mapa_cols_excepciones = {'Numero': COL_DNI, 'Nombre Completo': COL_NOMBRE_APELLIDO, 'Local': COL_LOCAL, 'Quien Autoriza': 'Quien_Autoriza', 'Fecha de Alta': COL_VENCE}
    df_excepciones, ult_mod_excepciones = cargar_y_procesar_excel(EXCEL_EXCEPCIONES, ult_mod_excepciones, 'Excepcion', mapa_cols_excepciones, df_excepciones, header=0, dni_col_original='Numero')
    if not df_excepciones.empty and COL_VENCE in df_excepciones.columns:
        df_excepciones[COL_VENCE] = pd.to_datetime(df_excepciones[COL_VENCE], errors='coerce')

    mapa_cols_nominas = {'DNI': COL_DNI, 'Apellido': 'Apellido', 'Nombre': 'Nombre', 'Empresa': COL_LOCAL, 'Vigencia Desde': 'Vigencia Desde', 'Vigencia Hasta': 'Vigencia Hasta'}
    df_nominas, ult_mod_nominas = cargar_y_procesar_excel(EXCEL_NOMINAS, ult_mod_nominas, 'Nomina', mapa_cols_nominas, df_nominas, header=0, dni_col_original='DNI')
    if not df_nominas.empty:
        if 'Nombre' in df_nominas.columns and 'Apellido' in df_nominas.columns:
            df_nominas[COL_NOMBRE_APELLIDO] = df_nominas['Nombre'].fillna('') + ' ' + df_nominas['Apellido'].fillna('')
        if 'Vigencia Desde' in df_nominas.columns:
            df_nominas['Vigencia Desde'] = pd.to_datetime(df_nominas['Vigencia Desde'], errors='coerce', dayfirst=True)
        if 'Vigencia Hasta' in df_nominas.columns:
            df_nominas['Vigencia Hasta'] = pd.to_datetime(df_nominas['Vigencia Hasta'], errors='coerce', dayfirst=True)

def get_nominas_agrupadas():
    """
    Lee el archivo de nóminas persistentes y devuelve un resumen agrupado
    por Empresa y rango de fechas.
    """
    try:
        cargar_autorizaciones()  # Asegura que los datos estén actualizados
        # Usamos la variable global df_nominas que ya está cargada y limpia
        df = df_nominas.copy()
        if df.empty:
            return []

        # Convertir fechas a string para agrupar
        df['Vigencia_Desde_str'] = df['Vigencia Desde'].dt.strftime('%d/%m/%Y')
        df['Vigencia_Hasta_str'] = df['Vigencia Hasta'].dt.strftime('%d/%m/%Y')
        df['Rango_Vigencia'] = df['Vigencia_Desde_str'] + ' - ' + df['Vigencia_Hasta_str']

        # Agrupar y contar
        agrupado = df.groupby([COL_LOCAL, 'Rango_Vigencia']).size().reset_index(name='Cantidad_Personas')
        
        # Convertir a formato JSON
        nominas_list = []
        for index, row in agrupado.iterrows():
            nominas_list.append({
                'id': f"{row[COL_LOCAL]}_{row['Rango_Vigencia']}", # Creamos un ID único
                'empresa': row[COL_LOCAL],
                'vigencia': row['Rango_Vigencia'],
                'cantidad_personas': row['Cantidad_Personas']
            })
        return nominas_list
        
    except Exception as e:
        print(f"Error al agrupar nóminas: {e}")
        return []  
def delete_nomina_by_criteria(empresa, vigencia):
    """
    Elimina registros del archivo de nóminas persistentes basados en la empresa
    y el string de vigencia (ej: "01/01/2025 - 31/12/2025").
    """
    try:
        df = df_nominas.copy()
        if df.empty:
            print("INFO: El DataFrame de nóminas en memoria está vacío. No hay nada que borrar.")
            return True

        # --- Lógica para comparar rangos de vigencia ---
        df['Vigencia_Desde_str'] = df['Vigencia Desde'].dt.strftime('%d/%m/%Y')
        df['Vigencia_Hasta_str'] = df['Vigencia Hasta'].dt.strftime('%d/%m/%Y')
        df['Rango_Vigencia'] = df['Vigencia_Desde_str'] + ' - ' + df['Vigencia_Hasta_str']
        # --- Fin lógica de rangos ---

        # Determinar qué columna de empresa usar para el filtrado
        columna_empresa_actual = COL_LOCAL if COL_LOCAL in df.columns else 'Empresa'
        if columna_empresa_actual not in df.columns:
            print(f"Error: No se encuentra la columna '{COL_LOCAL}' ni 'Empresa' en las nóminas.")
            return False

        # Condición para MANTENER registros que NO coincidan con los criterios de borrado
        condition = ~((df[columna_empresa_actual] == empresa) & (df['Rango_Vigencia'] == vigencia))
        df_filtrado = df[condition].copy()

        # --- Preparar para guardar ---
        # Renombrar la columna de empresa a 'Empresa' para consistencia en el archivo
        if columna_empresa_actual != 'Empresa' and columna_empresa_actual in df_filtrado.columns:
            df_filtrado.rename(columns={columna_empresa_actual: 'Empresa'}, inplace=True)

        # Definir las columnas que SIEMPRE deben estar en el archivo final
        columnas_finales = ['DNI', 'Apellido', 'Nombre', 'Empresa', 'Vigencia Desde', 'Vigencia Hasta']
        
        # Asegurarse de que el dataframe a guardar tenga todas las columnas, incluso si está vacío
        df_a_guardar = pd.DataFrame(columns=columnas_finales)
        
        # Si hay datos después de filtrar, usamos solo las columnas que nos interesan
        if not df_filtrado.empty:
            columnas_existentes = [col for col in columnas_finales if col in df_filtrado.columns]
            df_a_guardar = pd.concat([df_a_guardar, df_filtrado[columnas_existentes]], ignore_index=True)

        # Guardar el DataFrame filtrado de vuelta al archivo
        df_a_guardar.to_excel(EXCEL_NOMINAS, index=False)
        
        print(f"INFO: Nómina para '{empresa}' con vigencia {vigencia} eliminada.")
        
        # Forzar la recarga de datos en la memoria
        recargar_cache_nominas_persistentes()
        return True
        
    except Exception as e:
        print(f"Error Crítico al eliminar la nómina: {e}")
        return False

def update_nomina_entry(old_dni, old_empresa, old_vigencia_desde, old_vigencia_hasta,
                        new_dni, new_apellido, new_nombre, new_empresa, new_vigencia_desde, new_vigencia_hasta):
    """
    Modifica una entrada específica en el archivo de nóminas persistentes.
    Identifica la entrada por su DNI, Empresa y rango de vigencia antiguos,
    y la actualiza con los nuevos valores.
    """
    try:
        if not os.path.exists(EXCEL_NOMINAS):
            print("ADVERTENCIA: No existe archivo de nóminas para modificar.")
            return False

        df = pd.read_excel(EXCEL_NOMINAS)
        if df.empty:
            print("ADVERTENCIA: El archivo de nóminas está vacío, no hay nada que modificar.")
            return False

        # Asegurarse de que las columnas de fecha sean datetime para la comparación
        df['Vigencia Desde'] = pd.to_datetime(df['Vigencia Desde'], errors='coerce', dayfirst=True)
        df['Vigencia Hasta'] = pd.to_datetime(df['Vigencia Hasta'], errors='coerce', dayfirst=True)

        # Convertir los parámetros de fecha a datetime para la comparación
        old_vigencia_desde_dt = pd.to_datetime(old_vigencia_desde, errors='coerce', dayfirst=True)
        old_vigencia_hasta_dt = pd.to_datetime(old_vigencia_hasta, errors='coerce', dayfirst=True)

        # Identificar la fila a modificar
        # Usamos .astype(str) para DNI para asegurar la comparación si hay tipos mixtos
        # y .dt.normalize() para comparar solo la fecha sin la hora
        condicion = (
            (df['DNI'].astype(str) == str(old_dni)) &
            (df['Empresa'] == old_empresa) &
            (df['Vigencia Desde'].dt.normalize() == old_vigencia_desde_dt.normalize()) &
            (df['Vigencia Hasta'].dt.normalize() == old_vigencia_hasta_dt.normalize())
        )

        indices_a_modificar = df[condicion].index

        if indices_a_modificar.empty:
            print(f"ADVERTENCIA: No se encontró la entrada de nómina para modificar con DNI: {old_dni}, Empresa: {old_empresa}, Vigencia Desde: {old_vigencia_desde}, Vigencia Hasta: {old_vigencia_hasta}")
            return False

        # Aplicar las modificaciones
        for idx in indices_a_modificar:
            df.loc[idx, 'DNI'] = new_dni
            df.loc[idx, 'Apellido'] = new_apellido
            df.loc[idx, 'Nombre'] = new_nombre
            df.loc[idx, 'Empresa'] = new_empresa
            df.loc[idx, 'Vigencia Desde'] = pd.to_datetime(new_vigencia_desde, errors='coerce', dayfirst=True)
            df.loc[idx, 'Vigencia Hasta'] = pd.to_datetime(new_vigencia_hasta, errors='coerce', dayfirst=True)

        # Guardar el DataFrame modificado de vuelta al archivo
        df.to_excel(EXCEL_NOMINAS, index=False)

        print(f"INFO: Entrada de nómina modificada exitosamente para DNI: {old_dni}.")
        
        # Forzar la recarga de datos en la memoria
        cargar_autorizaciones()
        return True

    except Exception as e:
        print(f"Error Crítico al modificar la entrada de nómina: {e}")
        return False


def get_nomina_detalle_by_criteria(empresa, vigencia, filtro_dni=None, filtro_nombre=None, filtro_apellido=None):
    """
    Obtiene la lista detallada de una nómina y sus metadatos (fechas de vigencia),
    listos para ser usados en el formulario de edición.
    """
    try:
        df = df_nominas.copy()
        if df.empty:
            return None

        if 'Vigencia Desde' in df.columns and 'Vigencia Hasta' in df.columns:
            df['Vigencia Desde'] = pd.to_datetime(df['Vigencia Desde'], errors='coerce')
            df['Vigencia Hasta'] = pd.to_datetime(df['Vigencia Hasta'], errors='coerce')
            df['Vigencia_Desde_str'] = df['Vigencia Desde'].dt.strftime('%d/%m/%Y')
            df['Vigencia_Hasta_str'] = df['Vigencia Hasta'].dt.strftime('%d/%m/%Y')
            df['Rango_Vigencia'] = df['Vigencia_Desde_str'] + ' - ' + df['Vigencia_Hasta_str']
        else:
            return None

        df_filtrado = df[
            (df[COL_LOCAL] == empresa) &
            (df['Rango_Vigencia'] == vigencia)
        ].copy()

        if df_filtrado.empty:
            return None

        if filtro_dni:
            df_filtrado = df_filtrado[df_filtrado['DNI'].astype(str).str.contains(str(filtro_dni), case=False, na=False)]
        if filtro_nombre:
            df_filtrado = df_filtrado[df_filtrado['Nombre'].astype(str).str.contains(filtro_nombre, case=False, na=False)]
        if filtro_apellido:
            df_filtrado = df_filtrado[df_filtrado['Apellido'].astype(str).str.contains(filtro_apellido, case=False, na=False)]

        # Preparar la lista de personas con claves en minúscula
        df_personas = df_filtrado[['DNI', 'Apellido', 'Nombre']].copy()
        df_personas.rename(columns={'DNI': 'dni', 'Apellido': 'apellido', 'Nombre': 'nombre'}, inplace=True)
        
        # Extraer fechas y formatearlas para el input date (YYYY-MM-DD)
        vigencia_desde = df_filtrado['Vigencia Desde'].iloc[0].strftime('%Y-%m-%d')
        vigencia_hasta = df_filtrado['Vigencia Hasta'].iloc[0].strftime('%Y-%m-%d')

        return {
            'personas': df_personas.to_dict('records'),
            'vigencia_desde': vigencia_desde,
            'vigencia_hasta': vigencia_hasta
        }
        
    except Exception as e:
        print(f"Error al obtener detalle de nómina: {e}")
        return None

def get_single_nomina_entry(dni, empresa, vigencia_desde, vigencia_hasta):
    """
    Obtiene una única entrada de nómina para su edición.
    """
    try:
        df = df_nominas.copy()
        if df.empty:
            return None

        # Asegurarse de que las columnas de fecha sean datetime para la comparación
        df['Vigencia Desde'] = pd.to_datetime(df['Vigencia Desde'], errors='coerce', dayfirst=True)
        df['Vigencia Hasta'] = pd.to_datetime(df['Vigencia Hasta'], errors='coerce', dayfirst=True)

        # Convertir los parámetros de fecha a datetime para la comparación
        vigencia_desde_dt = pd.to_datetime(vigencia_desde, errors='coerce', dayfirst=True)
        vigencia_hasta_dt = pd.to_datetime(vigencia_hasta, errors='coerce', dayfirst=True)

        # Identificar la fila
        condicion = (
            (df['DNI'].astype(str) == str(dni)) &
            (df['Empresa'] == empresa) &
            (df['Vigencia Desde'].dt.normalize() == vigencia_desde_dt.normalize()) &
            (df['Vigencia Hasta'].dt.normalize() == vigencia_hasta_dt.normalize())
        )

        entry = df[condicion]

        if entry.empty:
            return None
        
        # Devolver la primera (y única) entrada encontrada como diccionario
        # Formatear las fechas a string para la respuesta JSON
        entry_dict = entry.iloc[0].to_dict()
        entry_dict['Vigencia Desde'] = entry_dict['Vigencia Desde'].strftime('%d/%m/%Y') if pd.notna(entry_dict['Vigencia Desde']) else None
        entry_dict['Vigencia Hasta'] = entry_dict['Vigencia Hasta'].strftime('%d/%m/%Y') if pd.notna(entry_dict['Vigencia Hasta']) else None
        
        return entry_dict

    except Exception as e:
        print(f"Error al obtener entrada de nómina individual: {e}")
        return None
    
def get_df_nominas_persistentes():
    """
    Devuelve el DataFrame de las nóminas persistentes que está en memoria.
    """
    global df_nominas
    return df_nominas

def recargar_cache_nominas_persistentes():
    """
    Alias para forzar la recarga de todos los DataFrames, incluidas las nóminas.
    """
    print("INFO: Forzando recarga de todos los cachés de autorización...")
    # Reseteamos los tiempos de modificación para forzar la recarga
    global ult_mod_fap, ult_mod_fao, ult_mod_excepciones, ult_mod_nominas
    ult_mod_fap, ult_mod_fao, ult_mod_excepciones, ult_mod_nominas = 0, 0, 0, 0
    
    cargar_autorizaciones()

def procesar_nomina_texto(texto_nomina):
    """
    Procesa un string multilínea que contiene una nómina de personal
    y la convierte en una lista de diccionarios.
    La lógica de parseo es robusta para manejar diferentes espaciados y formatos.
    """
    print("--- INICIANDO PROCESAMIENTO DE NÓMINA ---")
    lineas = texto_nomina.strip().splitlines() # Usar splitlines() para mejor manejo de saltos de línea
    personas_final = []

    # Expresiones regulares para varios formatos de línea.
    # Se procesan en orden. La primera que coincida se usa.
    formatos_regex = [
        # Formato 1: "C.U.I.L. 20408951853 ABALOS AXEL SEBASTIAN" (y variantes)
        (1, re.compile(r"^(?:C\.?U\.?I\.?L\.?[:\s]*)(?P<cuil>\d{11})\s+(?P<nombre_completo>.*)", re.IGNORECASE)),
        
        # Formato 2: "20-12345678-9 APELLIDO NOMBRE" (CUIL con guiones al inicio)
        (2, re.compile(r"^(?P<cuil>\d{2}-\d{7,8}-\d{1})\s+(?P<nombre_completo>.*)", re.IGNORECASE)),

        # Formato 3: "20365205044 11 FIGUEROA WALTER..." (CUIL, un número, y nombre)
        (3, re.compile(r"^(?P<cuil>\d{11})\s+\d+\s+(?P<nombre_completo>.*)", re.IGNORECASE)),

        # Formato 4: "20365205044 FIGUEROA WALTER..." (CUIL y nombre)
        (4, re.compile(r"^(?P<cuil>\d{11})\s+(?P<nombre_completo>.*)", re.IGNORECASE)),

        # Formato 5: DNI (7-8 digitos) y nombre
        (5, re.compile(r"^(?P<dni>\d{7,8})\s+(?P<nombre_completo>.*)", re.IGNORECASE)),

        # Formato 6: "APELLIDO NOMBRE 20-12345678-9" (CUIL con guiones al final)
        (6, re.compile(r"^(?P<nombre_completo>.*?)\s+(?P<cuil>\d{2}-\d{7,8}-\d{1})$", re.IGNORECASE)),
    ]

    print(f"Texto recibido para procesar:\n---\n{texto_nomina}\n---")
    print(f"Procesando {len(lineas)} líneas.")

    for i, linea in enumerate(lineas):
        linea = linea.strip()
        print(f"\n[Línea {i+1}]: '{linea}'")
        if not linea:
            print("-> Línea vacía, ignorando.")
            continue

        # Ignorar encabezados comunes
        if any(h in linea.upper() for h in ['CUIL', 'APELLIDO', 'NOMBRE', 'LEGAJO', 'LEGAJOS']):
            print("-> Línea parece un encabezado, ignorando.")
            continue

        match = None
        matched_format = 0
        for fmt, regex in formatos_regex:
            match = regex.match(linea)
            if match:
                matched_format = fmt
                break
        
        if not match:
            print(f"-> ADVERTENCIA: La línea no coincide con ningún formato conocido.")
            continue

        print(f"-> Coincide con formato #{matched_format}.")

        try:
            datos = match.groupdict()
            print(f"   - Datos extraídos: {datos}")
            dni = ""
            apellido = ""
            nombre = ""

            if 'cuil' in datos:
                dni = extraer_dni_de_cuil(datos['cuil'])
                print(f"   - CUIL '{datos['cuil']}' -> DNI '{dni}'")
            elif 'dni' in datos:
                dni = datos['dni']
                print(f"   - DNI encontrado: '{dni}'")

            nombre_completo_str = datos.get('nombre_completo', '').strip()
            
            # Eliminar "Régimen General" si está al final
            if nombre_completo_str.lower().endswith("régimen general"):
                nombre_completo_str = nombre_completo_str[:-15].strip()

            if not nombre_completo_str:
                print(f"-> ADVERTENCIA: No se pudo extraer el nombre completo.")
                continue

            print(f"   - Nombre completo a procesar: '{nombre_completo_str}'")
            partes_nombre = nombre_completo_str.split()
            
            # Lógica para separar Apellido y Nombre
            if len(partes_nombre) >= 2:
                # Asume que la primera palabra es el apellido y el resto es el nombre.
                apellido = partes_nombre[0]
                nombre = ' '.join(partes_nombre[1:])
            elif len(partes_nombre) == 1:
                # Si solo hay una palabra, se considera apellido.
                apellido = partes_nombre[0]
                nombre = ""
            
            print(f"   - Apellido: '{apellido}', Nombre: '{nombre}'")

            if dni and (nombre or apellido):
                persona = {'DNI': dni, 'Apellido': apellido, 'Nombre': nombre}
                personas_final.append(persona)
                print(f"-> ÉXITO: Persona agregada: {persona}")
            else:
                print(f"-> ADVERTENCIA: No se pudo extraer DNI o Nombre/Apellido válido.")

        except (ValueError, IndexError) as e:
            print(f"-> ERROR: La línea no pudo ser procesada, error: {e}.")
            continue
            
    print(f"--- PROCESAMIENTO FINALIZADO: {len(personas_final)} personas encontradas. ---\n")
    return personas_final

def generar_reporte_consolidado():
    """
    Lee el registro diario, consolida las entradas y salidas en una sola fila por DNI,
    y devuelve la ruta a un archivo Excel temporal con el reporte.
    """
    fecha_actual_str = datetime.now().strftime('%Y-%m-%d')
    nombre_archivo_original = os.path.join(REGISTROS_DIARIOS_DIR, f'registros_ingreso_{fecha_actual_str}.xlsx')

    if not os.path.exists(nombre_archivo_original):
        return None, None

    df = pd.read_excel(nombre_archivo_original)
    
    df['DNI'] = df['DNI'].astype(str)
    df['Hora_Ingreso'] = df['Hora_Ingreso'].astype(str).replace(['NaT', 'nan'], '')
    df['Hora_Salida'] = df['Hora_Salida'].astype(str).replace(['NaT', 'nan'], '')

    consolidado_df = df.groupby('DNI').agg({
        'Nombre y Apellido': 'first',
        'Hora_Ingreso': 'first',
        'Hora_Salida': 'last',
        'Evento': 'first',
        'Tipo_Permiso': 'first',
        'Num_Permiso': 'first',
        'Local': 'first',
        'Tarea': 'first',
        'Resultado': 'first' 
    }).reset_index()

    def calcular_resultado_final(row):
        if row['Resultado'] == 'DENEGADO':
            return 'DENEGADO'
        if row['Hora_Salida'] and row['Hora_Salida'] != '':
            return 'Registrado'
        if row['Hora_Ingreso'] and row['Hora_Ingreso'] != '':
            return 'Autorizado'
        return row['Resultado']

    consolidado_df['Resultado'] = consolidado_df.apply(calcular_resultado_final, axis=1)
    
    columnas_finales = ['DNI', 'Nombre y Apellido', 'Hora_Ingreso', 'Hora_Salida', 'Evento', 'Tipo_Permiso', 'Num_Permiso', 'Local', 'Tarea', 'Resultado']
    consolidado_df = consolidado_df.reindex(columns=columnas_finales)

    temp_dir = os.path.join(BASE_DIR, 'temp')
    os.makedirs(temp_dir, exist_ok=True)
    nombre_archivo_temp = os.path.join(temp_dir, f'reporte_consolidado_{fecha_actual_str}.xlsx')
    
    consolidado_df.to_excel(nombre_archivo_temp, index=False)
    
    formatear_excel(nombre_archivo_temp)

    return nombre_archivo_temp, os.path.basename(nombre_archivo_temp)            
    